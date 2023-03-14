import os, pandas as pd
from openpyxl import load_workbook
from Analysis_Pipeline import importData

'''This script rewrites the entire All Data data table. Any custom entries and formatting will be wiped. Use with caution'''



# Clears any remaining data out of the data table - only performed at the start of the program
wb = load_workbook("Migration Speed Analysis.xlsx")
main_table = wb['All Data']

main_table.delete_rows(1, main_table.max_row)
main_table.delete_cols(1, main_table.max_column)
wb.save('Migration Speed Analysis.xlsx')


# Defines all directories that contain program mapping spreadsheets to be referenced
myDirectory = os.getcwd()
odysseyDirectory = os.path.abspath(os.path.join(myDirectory, os.pardir))
mappingsDirectory = odysseyDirectory + '\Program Mapping Spreadsheets'
completedMappingsDir = mappingsDirectory + '/Completed Migrations'
oldMappingsDir = completedMappingsDir + '/Migrations pre-2023'


# Uses the file name to find the name of the agency based on the typical formatting of the file names
def findAgencyName(file):
    # Split the file name so it can be read word-by-word
    filename = file.split()
    agency = ''
    
    # Certain older files are formatted with the date first, this if statement checks for those
    if filename[0] == 'Migrate':
        index = 2

        # If the current word is not a number, append it to the agency name. Keep running this loop until a number is found
        while not any(char.isdigit() for char in filename[index]):
            agency += filename[index] + ' '
            index += 1

    else:
        index = 0

        # If the current word is not a number, append it to the agency name. Keep running this loop until a number is found
        while not any(char.isdigit() for char in filename[index]):
            agency += filename[index] + ' '
            index += 1

    return agency.strip()



# Function to check whether 'results - final migration' sheet exists in the file
def data_exists(filepath):
    excel_file = pd.ExcelFile(filepath)
    sheet_to_find = "results - final migration"

    # Get all the sheetnames as a list
    sheet_names = excel_file.sheet_names

    # Format the list of sheet names
    sheet_names = [name.lower() for name in sheet_names]

    # Check whether 'results - final migration' is the name of any of the file's sheets
    if sheet_to_find in sheet_names:
        return True
    return False


# Main Function
def importFromDirectory(dir):
    # Iterates over each file in the specified folder
    for file in os.listdir(dir):

        # Ensure that the current file is an Excel file
        if not str(file).endswith('.xlsx'):
            continue

        # Format into a full directory for Python to search into
        agency = findAgencyName(file)
        filepath = dir + '/' + file
        # Print the agency for each iteration for QA purposes
        print(agency)

        #Ensures that the 'results - final migration' tab exists within the file before letting the import script search for it
        if not data_exists(filepath):
            print('Skipped')
            continue

        importData(agency, filepath)



importFromDirectory(oldMappingsDir)
importFromDirectory(completedMappingsDir)
importFromDirectory(mappingsDirectory)