from openpyxl import load_workbook

'''PSA: This script is made mainly for testing and should not have to be run regularly'''

# This is the only input needed, script will find rows associated with this agency and delete them
agency = 'Gracewood'


# Opens analysis file and the main data sheet specifically
wb = load_workbook("Migration Speed Analysis copy.xlsx")
main_table = wb['All Data']

# Amount of rows deleted is tracked for quality assurance
deletedCount = 0

# Openpyxl starts its indexing at 1 instead of 0
index = 1

# If a row is deleted and the index goes up, a row is skipped. In order to have a non-linear
# index, a sort of manual iteration is done. The index is defined outside of the function
# and increased only when a row is not deleted.
while index <= main_table.max_row:
    if main_table[index][0].value == agency:
        main_table.delete_rows(index)
        deletedCount += 1
    else:
        index += 1

wb.save("Migration Speed Analysis copy.xlsx")
print('Deleted', deletedCount, 'rows from data file.')