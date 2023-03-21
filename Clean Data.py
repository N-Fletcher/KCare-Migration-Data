from openpyxl import load_workbook

dataFile = "Migration Speed Analysis.xlsx"

# Define all of the sheets being accessed and/or written into
wb = load_workbook(dataFile)
raw = wb['All Data']
clean = wb['All Data - Clean']
old = wb['Past Data']
new = wb['2023 Data']


# Empty all of the sheets that are getting written into
clean.delete_rows(2, clean.max_row)
old.delete_rows(2, old.max_row)
new.delete_rows(2, new.max_row)
wb.save(dataFile)


# Iterate over all-encompassing data table
for row in raw:

    # Define each row's necessary contents for readability
    date = row[1].value
    migrated = row[8].value
    time = float(row[10].value)

    # Entries with miniscule amounts of data skew the 'items per minute' averages, so they are skipped over automatically
    if time < 2.0: continue
    if migrated == 0: continue

    # Convert the current row's contents into a list, so that it can be read into workbook's append statement
    currentRow = [item.value for item in row]

    # Append all remaining data to the All Data - Clean' tab
    clean.append(currentRow)
    
    # Separate remaining data by date
    if date.__contains__('2023'):
        new.append(currentRow)
    else:
        old.append(currentRow)


wb.save(dataFile)