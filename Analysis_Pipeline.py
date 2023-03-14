import pandas as pd, re
from openpyxl import load_workbook



'''External Functions'''

# Function cleans up 'Took' column so that all values can be read as integers
def formatTime(value) -> float:

    if value.strip().__contains__("<1 minute") | value.strip().__contains__("< 1 minute"):
        return 0.5
    return value.split()[0]


# Function calculates updated, new and migrated values, accounting for commas
def calculate(updated, new) -> tuple[int, int, int]:

    # If statements account for values with commas (values in the thousands)
    if updated.__contains__(','):
        updated = int(updated.strip().replace(',', ''))
    else:
        updated = int(updated)

    if new.__contains__(','):
        new = int(new.strip().replace(',', ''))
    else:
        new = int(new)

    migrated = updated + new
    return updated, new, migrated


# Function appends a given row to the main data file
def appendRow(output_file, agency, date, system, category, document, total, new,
              updated, migrated, skipped, time, speed):
    newRow = [agency, date, system, category, document, total, new,
                         updated, migrated, skipped, time, speed]
    
    wb = load_workbook(output_file)
    ws = wb.get_sheet_by_name('All Data')

    ws.append(newRow)
    wb.save(output_file)


def findSheet(inputFile):
    excel_file = pd.ExcelFile(inputFile)

    # Get all the sheetnames as a list
    sheet_names = excel_file.sheet_names

    # Format the list of sheet names
    sheet_names = [name.lower() for name in sheet_names]

    # Get the index that matches our sheet to find
    return sheet_names.index('results - final migration')


'''Main function'''
def importData(agency, input_file, output_file="Migration Speed Analysis.xlsx"):

    input = pd.read_excel(input_file, sheet_name=findSheet(input_file))
    #pandas automatically sets empty cells as 'NaN', but the code was built to work with empty strings, so empty cells are manually set to ''
    input.fillna('', inplace=True)


    # script tracks how many rows are appended for quality assurance
    newRowsCount = 0
    # Date should be a static value, so it's defined outside of the scope of loops
    date = ''

    for _, row in input.iterrows():

        # There's only one date for an entire migration, so the code first checks
        # whether the date was already found. If not, it keeps checking each row
        # until it finds the date
        if date == '':
            temp = str(row[3])
            if temp.__contains__('/') | temp.__contains__('-'):
                date = temp.split()[0]

        # Constantly checks for header box that contains all of the migration info,
        # looks for 'FC' and 'GCM' keywords within that box
        if str(row[0]).lower().__contains__('(fc') | str(row[0]).lower().__contains__('fc/'):
            system = 'FC'               # script finds unintended 'fc' characters, '(' combats this

        elif str(row[0]).lower().__contains__('gcm'):
            system = 'GCM'

        # Checks to see whether row contains data first, then runs transfer code
        time_column = str(row[8])
        if (row[2] != '') & (time_column.__contains__('minute') | time_column.strip().isdigit()):

            #Identify any '-' symbols and assume they should be read as 0's
            for index in range(8):
                if row[index] == '-':
                    row[index] = 0

            # Column mappings specifically outlined to avoid headaches
            # later on should formatting change at all
            document = row[2]
            category = row[0]
            total = row[4]
            updated, new, migrated = calculate(str(row[5]), str(row[6]))
            skipped = row[7]

            time = formatTime(row[8])
            if time == '0':
                continue

            speed = float(migrated) / float(time)

            # writing function externalized to avoid quadruple nesting
            appendRow(output_file, agency, date, system, category, document, total, new,
                    updated, migrated, skipped, time, speed)

            newRowsCount += 1
    print('Appended', newRowsCount, 'rows to data file.')